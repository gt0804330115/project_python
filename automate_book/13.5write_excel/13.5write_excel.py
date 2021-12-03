"""
    写入Excel文档
"""
# 创建并保存Excel文档
# import openpyxl
# wb = openpyxl.Workbook()
# print(wb.sheetnames)
# sheet = wb.active
# print(sheet.title)
# sheet.title = 'Span Bacon Eggs Sheet'
# print(wb.sheetnames)
# wb.save('example_copy.xlsx')

# 创建和删除工作表
import openpyxl

wb = openpyxl.Workbook()
wb.create_sheet()  # 添加一个新的工作表
print(wb.sheetnames)
wb.create_sheet(index=0, title='First sheet')
print(wb.sheetnames)
wb.create_sheet(index=2, title='Middle sheet')
print(wb.sheetnames)
del wb['Middle sheet']
del wb['Sheet1']
print(wb.sheetnames)


# 将值写入单元格
import openpyxl
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Hello world!'
print(sheet['A1'].value)
