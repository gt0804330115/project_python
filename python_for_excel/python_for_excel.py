"""
    Excel电子表格操作
"""
import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))  # <class 'openpyxl.workbook.workbook.Workbook'>

# 从工作簿中取得工作表
sheet = wb['Sheet3']
print(sheet)  # <Worksheet "Sheet3">
print(sheet.title)  # Sheet3

antherSheet = wb.active  # active属性  -->  获取工作簿的活动表
print(antherSheet)  # <Worksheet "Sheet1">

# 从表中取得单元格
sheet = wb['Sheet1']
print(sheet['A1'])  # <Cell 'Sheet1'.A1>
print(sheet['A1'].value)  # 4/5/2015 1:34:02 PM
c = sheet['B1']
print(c.value)  # Apples
print('Row %s,Column %s is %s' % (c.row, c.column, c.value))  # Row 1,Column 2 is Apples
print('Cell %s is %s' % (c.coordinate, c.value))  # Cell B1 is Apples
print(sheet['C1'].value)  # 73
# row --> 行 ， column --> 列
print(sheet.cell(row=1, column=2))  # <Cell 'Sheet1'.B1>
print(sheet.cell(row=1, column=2).value)  # Apples
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
    # 1 Apples
    # 3 Pears
    # 5 Apples
    # 7 Strawberries
# 获取表格中最大行的值--> max_row
# 获取表格中最大列的值--> max_column
print(sheet.max_row)  # 7
print(sheet.max_column)  # 3
# 列字母和数字之间的转换
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

print(get_column_letter(1))  # A
print(get_column_letter(2))  # B
print(get_column_letter(27))  # AA
print(get_column_letter(900))  # AHP
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
print(get_column_letter(sheet.max_column))  # C
print(column_index_from_string('A'))  # 1
print(column_index_from_string('AA'))  # 27

# 从表中取得行和列
import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
print(tuple(sheet[
            'A1':'C3']))  # ((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>), (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>), (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>))
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('---END OF ROW---')
    # A1 4/5/2015 1:34:02 PM
    # B1 Apples
    # C1 73
    # ---END OF ROW---
    # A2 4/5/2015 3:41:23 AM
    # B2 Cherries
    # C2 85
    # ---END OF ROW---
    # A3 4/6/2015 12:46:51 PM
    # B3 Pears
    # C3 14
    # ---END OF ROW---
import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
print(list(sheet.columns)[1])  # (<Cell 'Sheet1'.B1>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.B4>, <Cell 'Sheet1'.B5>, <Cell 'Sheet1'.B6>, <Cell 'Sheet1'.B7>)
for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)
    # Apples
    # Cherries
    # Pears
    # Oranges
    # Apples
    # Bananas
    # Strawberries
